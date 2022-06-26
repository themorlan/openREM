# This Python file uses the following encoding: utf-8
# test_export_ct_xlsx.py

import os
from django.contrib.auth.models import User, Group
from django.test import TestCase, RequestFactory
from openpyxl import load_workbook
from remapp.extractors import rdsr
from remapp.exports.ct_export import ctxlsx, ct_csv, ct_phe_2019
from remapp.models import PatientIDSettings, Exports


class ExportCTxlsx(TestCase):
    """Test class for CT exports to XLSX"""

    def setUp(self):
        self.factory = RequestFactory()
        self.user = User.objects.create_user(
            username="jacob", email="jacob@…", password="top_secret"
        )
        eg = Group(name="exportgroup")
        eg.save()
        eg.user_set.add(self.user)
        eg.save()

        pid = PatientIDSettings.objects.create()
        pid.name_stored = True
        pid.name_hashed = False
        pid.id_stored = True
        pid.id_hashed = False
        pid.dob_stored = True
        pid.save()

        ct_ge_ct660 = os.path.join("test_files", "CT-ESR-GE_Optima.dcm")
        ct_ge_vct = os.path.join("test_files", "CT-ESR-GE_VCT.dcm")
        ct_siemens_flash_ss = os.path.join(
            "test_files", "CT-RDSR-Siemens_Flash-TAP-SS.dcm"
        )
        ct_toshiba_dosecheck = os.path.join(
            "test_files", "CT-RDSR-Toshiba_DoseCheck.dcm"
        )
        root_tests = os.path.dirname(os.path.abspath(__file__))

        rdsr.rdsr(os.path.join(root_tests, ct_ge_ct660))
        rdsr.rdsr(os.path.join(root_tests, ct_ge_vct))
        rdsr.rdsr(os.path.join(root_tests, ct_siemens_flash_ss))
        rdsr.rdsr(os.path.join(root_tests, ct_toshiba_dosecheck))

    def test_id_as_text(self):  # See https://bitbucket.org/openrem/openrem/issues/443
        filter_set = {"o": "-study_date"}
        pid = True
        name = False
        patient_id = True

        ctxlsx(filter_set, pid=pid, name=name, patid=patient_id, user=self.user)

        task = Exports.objects.all()[0]

        book = load_workbook(task.filename.path)
        all_data_sheet = book["All data"]
        headers = all_data_sheet[1]

        patient_id_col = [
            i for i, x in enumerate(headers, start=1) if x.value == "Patient ID"
        ][0]
        accession_number_col = [
            i for i, x in enumerate(headers, start=1) if x.value == "Accession number"
        ][0]
        dlp_total_col = [
            i for i, x in enumerate(headers, start=1) if x.value == "DLP total (mGy.cm)"
        ][0]
        e1_dose_check_col = [
            i
            for i, x in enumerate(headers, start=1)
            if x.value == "E1 Dose check details"
        ][0]
        e2_dose_check_col = [
            i
            for i, x in enumerate(headers, start=1)
            if x.value == "E2 Dose check details"
        ][0]

        self.assertEqual(
            all_data_sheet.cell(row=3, column=patient_id_col).data_type, "s"
        )
        self.assertEqual(
            all_data_sheet.cell(row=4, column=patient_id_col).data_type, "s"
        )
        self.assertEqual(
            all_data_sheet.cell(row=3, column=accession_number_col).data_type, "s"
        )
        self.assertEqual(
            all_data_sheet.cell(row=4, column=accession_number_col).data_type, "s"
        )
        self.assertEqual(
            all_data_sheet.cell(row=3, column=dlp_total_col).data_type, "n"
        )

        self.assertEqual(
            all_data_sheet.cell(row=3, column=patient_id_col).value, "008F/g234"
        )
        self.assertEqual(
            all_data_sheet.cell(row=4, column=patient_id_col).value, "00001234"
        )
        self.assertEqual(
            all_data_sheet.cell(row=3, column=accession_number_col).value,
            "001234512345678",
        )
        self.assertEqual(
            all_data_sheet.cell(row=4, column=accession_number_col).value,
            "0012345.12345678",
        )
        self.assertEqual(
            all_data_sheet.cell(row=3, column=dlp_total_col).value, 2002.39
        )

        e1_dose_check_string = (
            "Dose Check Alerts: DLP alert is configured at 100.00 mGy.cm with an accumulated "
            "forward estimate of 251.20 mGy.cm. CTDIvol alert is configured at 10.00 mGy with no "
            "accumulated forward estimate recorded. Person authorizing irradiation: Luuk. "
        )
        e2_dose_check_string = (
            "Dose Check Alerts: DLP alert is configured at 100.00 mGy.cm with an accumulated "
            "forward estimate of 502.40 mGy.cm. CTDIvol alert is configured at 10.00 mGy with an "
            "accumulated forward estimate of 10.60 mGy. Person authorizing irradiation: Luuk. "
        )
        self.assertEqual(
            all_data_sheet.cell(row=2, column=e1_dose_check_col).value,
            e1_dose_check_string,
        )
        self.assertEqual(
            all_data_sheet.cell(row=2, column=e2_dose_check_col).value,
            e2_dose_check_string,
        )

        # cleanup
        task.filename.delete()  # delete file so local testing doesn't get too messy!

    def test_zero_filter(self):
        """Test error handled correctly when empty filter."""
        filter_set = {"study_description": "asd"}
        pid = True
        name = False
        patient_id = True

        ctxlsx(filter_set, pid=pid, name=name, patid=patient_id, user=self.user)

        task = Exports.objects.all()[0]
        self.assertEqual("ERROR", task.status)

    def test_acq_type_filter_spiral(self):
        """Test to check that filtering CT by acquisition type works
        as expected.

        """
        filter_set = {
            "num_spiral_events": "some",
            "o": "-study_date",
        }
        pid = True
        name = False
        patient_id = True

        ctxlsx(filter_set, pid=pid, name=name, patid=patient_id, user=self.user)

        task = Exports.objects.all()[0]
        self.assertEqual(4, task.num_records)

        # cleanup
        task.filename.delete()  # delete file so local testing doesn't get too messy!

    def test_acq_type_filter_sequenced(self):
        """Test to check that filtering CT by acquisition type works
        as expected.

        """
        filter_set = {
            "num_axial_events": "some",
            "o": "-study_date",
        }
        pid = True
        name = False
        patient_id = True

        ctxlsx(filter_set, pid=pid, name=name, patid=patient_id, user=self.user)

        task = Exports.objects.all()[0]
        self.assertEqual(1, task.num_records)

        # cleanup
        task.filename.delete()  # delete file so local testing doesn't get too messy!

    def test_acq_type_filter_spiral_and_sequenced(self):
        """Test to check that filtering CT by acquisition type works
        as expected.

        """
        filter_set = {
            "num_spiral_events": "some",
            "num_axial_events": "some",
            "o": "-study_date",
        }
        pid = True
        name = False
        patient_id = True

        ctxlsx(filter_set, pid=pid, name=name, patid=patient_id, user=self.user)

        task = Exports.objects.all()[0]
        self.assertEqual(1, task.num_records)

        # cleanup
        task.filename.delete()  # delete file so local testing doesn't get too messy!

    def test_export_phe(self):
        filter_set = {"num_spiral_events": "2", "o": "-study_date"}

        ct_phe_2019(filter_set, user=self.user)

        task = Exports.objects.order_by("pk")[0]
        self.assertEqual(3, task.num_records)

        book = load_workbook(task.filename.path)
        phe_sheet = book["PHE CT 2019"]

        self.assertEqual(
            phe_sheet.cell(column=5, row=2).value, 487
        )  # first series imaged length
        self.assertEqual(phe_sheet.cell(row=3, column=5).value, 5)
        self.assertEqual(phe_sheet.cell(row=4, column=5).value, 418.75)
        self.assertEqual(
            phe_sheet.cell(row=2, column=11).value, 5.3
        )  # first series CTDIvol
        self.assertEqual(phe_sheet.cell(row=3, column=11).value, 32.83)
        self.assertEqual(phe_sheet.cell(row=4, column=11).value, 3.23)
        self.assertEqual(
            phe_sheet.cell(row=2, column=20).value, 251.2
        )  # second series DLP
        self.assertEqual(phe_sheet.cell(row=3, column=20).value, 429.19)
        self.assertEqual(phe_sheet.cell(row=4, column=20).value, 259.85)
        self.assertEqual(phe_sheet.cell(row=2, column=37).value, 502.4)  # total DLP
        self.assertEqual(phe_sheet.cell(row=3, column=37).value, 2002.39)
        self.assertEqual(phe_sheet.cell(row=4, column=37).value, 415.82)
        self.assertEqual(phe_sheet.cell(row=4, column=3).value, 78.2)
        self.assertEqual(phe_sheet.cell(row=4, column=4).value, 164)

        # cleanup
        task.filename.delete()  # delete file so local testing doesn't get too messy!
