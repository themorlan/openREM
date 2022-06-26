# This Python file uses the following encoding: utf-8
# test_export_dx_xlsx.py

import hashlib
import os

from django.contrib.auth.models import User, Group
from django.test import TestCase, RequestFactory
from openpyxl import load_workbook

from remapp.extractors import dx
from remapp.exports.dx_export import dxxlsx
from remapp.models import PatientIDSettings, Exports


class ExportDXxlsx(TestCase):
    def setUp(self):
        self.factory = RequestFactory()
        self.user = User.objects.create_user(
            username="jacob", email="jacob@…", password="top_secret"
        )
        eg = Group(name="exportgroup")
        eg.save()
        eg.user_set.add(self.user)
        eg.save()

        pid = PatientIDSettings.objects.create()
        pid.name_stored = True
        pid.name_hashed = False
        pid.id_stored = True
        pid.id_hashed = False
        pid.dob_stored = True
        pid.save()

        dx_ge_xr220_1 = os.path.join("test_files", "DX-Im-GE_XR220-1.dcm")
        dx_ge_xr220_2 = os.path.join("test_files", "DX-Im-GE_XR220-2.dcm")
        dx_ge_xr220_3 = os.path.join("test_files", "DX-Im-GE_XR220-3.dcm")
        dx_carestream_dr7500_1 = os.path.join(
            "test_files", "DX-Im-Carestream_DR7500-1.dcm"
        )
        dx_carestream_dr7500_2 = os.path.join(
            "test_files", "DX-Im-Carestream_DR7500-2.dcm"
        )
        root_tests = os.path.dirname(os.path.abspath(__file__))

        dx.dx(os.path.join(root_tests, dx_ge_xr220_1))
        dx.dx(os.path.join(root_tests, dx_ge_xr220_2))
        dx.dx(os.path.join(root_tests, dx_ge_xr220_3))
        dx.dx(os.path.join(root_tests, dx_carestream_dr7500_1))
        dx.dx(os.path.join(root_tests, dx_carestream_dr7500_2))

    def test_id_as_text(self):  # See https://bitbucket.org/openrem/openrem/issues/443
        filter_set = {}
        pid = True
        name = False
        patient_id = True

        dxxlsx(filter_set, pid=pid, name=name, patid=patient_id, user=self.user)

        task = Exports.objects.order_by("task_id")[0]

        book = load_workbook(task.filename.path)
        all_data_sheet = book["All data"]
        headers = all_data_sheet[1]

        patient_id_col = [
            i for i, x in enumerate(headers, start=1) if x.value == "Patient ID"
        ][0]
        accession_number_col = [
            i for i, x in enumerate(headers, start=1) if x.value == "Accession number"
        ][0]
        exposure_index_col = [
            i for i, x in enumerate(headers, start=1) if x.value == "E1 Exposure index"
        ][0]

        self.assertEqual(
            all_data_sheet.cell(row=2, column=patient_id_col).data_type, "s"
        )
        self.assertEqual(
            all_data_sheet.cell(row=2, column=accession_number_col).data_type, "s"
        )
        self.assertEqual(
            all_data_sheet.cell(row=2, column=exposure_index_col).data_type, "n"
        )

        self.assertEqual(
            all_data_sheet.cell(row=2, column=patient_id_col).value, "00098765"
        )
        self.assertEqual(
            all_data_sheet.cell(row=2, column=accession_number_col).value, "00938475"
        )
        self.assertEqual(
            all_data_sheet.cell(row=2, column=exposure_index_col).value, 51.745061
        )

        # cleanup
        task.filename.delete()  # delete file so local testing doesn't get too messy!
        task.delete()  # not necessary, by hey, why not?

    def test_filters(self):
        # Tests that extracts with multiple filters succeed (though previous test would fail too!
        filter_set = {}
        # filter_set = "display_name=Carestream+Clinic+KODAK7500&"
        pid = True
        name = True
        patient_id = True

        dxxlsx(filter_set, pid=pid, name=name, patid=patient_id, user=self.user)

        task = Exports.objects.order_by("task_id")[0]

        book = load_workbook(task.filename.path)
        aec_sheet = book["aec"]
        headers = aec_sheet[1]

        filter_col = [
            i for i, x in enumerate(headers, start=1) if x.value == "Filters"
        ][0]
        filter_thick_col = [
            i
            for i, x in enumerate(headers, start=1)
            if x.value == "Filter thicknesses (mm)"
        ][0]

        self.assertEqual(aec_sheet.cell(row=2, column=filter_col).value, "Al")
        self.assertEqual(aec_sheet.cell(row=2, column=filter_thick_col).value, "1.0000")
        self.assertEqual(aec_sheet.cell(row=3, column=filter_col).value, "Al | Cu")
        self.assertEqual(
            aec_sheet.cell(row=3, column=filter_thick_col).value, "1.0000 | 0.2000"
        )

        # cleanup
        task.filename.delete()  # delete file so local testing doesn't get too messy!
        task.delete()  # not necessary, by hey, why not?
