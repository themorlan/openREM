# This Python file uses the following encoding: utf-8
# test_export_rf_xlsx.py

import os

from django.contrib.auth.models import User, Group
from django.test import RequestFactory, TransactionTestCase
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from ..extractors import rdsr
from ..exports.rf_export import rfxlsx
from ..models import PatientIDSettings, Exports, GeneralStudyModuleAttr


class ExportRFxlsx(
    TransactionTestCase
):  # Not TestCase as raises TransactionManagementError
    def setUp(self):
        self.factory = RequestFactory()
        self.user = User.objects.create_user(
            username="jacob", email="jacob@…", password="top_secret"
        )
        eg = Group(name="exportgroup")
        eg.save()
        eg.user_set.add(self.user)
        eg.save()

        pid = PatientIDSettings.objects.create()
        pid.name_stored = True
        pid.name_hashed = False
        pid.id_stored = True
        pid.id_hashed = False
        pid.dob_stored = True
        pid.save()

        rf_siemens_zee = os.path.join("test_files", "RF-RDSR-Siemens-Zee.dcm")
        rf_philips_allura = os.path.join("test_files", "RF-RDSR-Philips_Allura.dcm")
        rf_eurocolumbus = os.path.join("test_files", "RF-RDSR-Eurocolumbus.dcm")
        root_tests = os.path.dirname(os.path.abspath(__file__))

        rdsr.rdsr(os.path.join(root_tests, rf_siemens_zee))
        rdsr.rdsr(os.path.join(root_tests, rf_philips_allura))
        rdsr.rdsr(os.path.join(root_tests, rf_eurocolumbus))

        eurocolumbus = GeneralStudyModuleAttr.objects.filter(
            study_instance_uid__exact="1.3.6.1.4.1.5962.99.1.1227319599.741127153.1517350807855.3.0"
        ).first()
        eurocolumbus.modality_type = "RF"
        eurocolumbus.save()

    def test_id_as_text(self):  # See https://bitbucket.org/openrem/openrem/issues/443
        filter_set = {"o": "-study_date"}
        pid = True
        name = False
        patient_id = True

        rfxlsx(filter_set, pid=pid, name=name, patid=patient_id, user=self.user)

        task = Exports.objects.all()[0]

        book = load_workbook(task.filename.path)
        all_data_sheet = book["All data"]
        headers = all_data_sheet[1]

        patient_id_col = [
            i for i, x in enumerate(headers, start=1) if x.value == "Patient ID"
        ][0]
        accession_number_col = [
            i for i, x in enumerate(headers, start=1) if x.value == "Accession number"
        ][0]
        a_dose_rp_col = [
            i
            for i, x in enumerate(headers, start=1)
            if x.value == "A Dose RP total (Gy)"
        ][0]
        manufacturer_col = [
            i for i, x in enumerate(headers, start=1) if x.value == "Manufacturer"
        ][0]
        manufacturers = all_data_sheet[get_column_letter(manufacturer_col)]
        siemens_row = [
            i for i, x in enumerate(manufacturers, start=1) if x.value == "Siemens"
        ][0]

        self.assertEqual(
            all_data_sheet.cell(row=siemens_row, column=patient_id_col).data_type, "s"
        )
        self.assertEqual(
            all_data_sheet.cell(row=siemens_row, column=accession_number_col).data_type,
            "s",
        )
        self.assertEqual(
            all_data_sheet.cell(row=siemens_row, column=a_dose_rp_col).data_type, "n"
        )

        self.assertEqual(
            all_data_sheet.cell(row=siemens_row, column=patient_id_col).value, "098765"
        )
        self.assertEqual(
            all_data_sheet.cell(row=siemens_row, column=accession_number_col).value,
            "1234.5678",
        )
        self.assertEqual(
            all_data_sheet.cell(row=siemens_row, column=a_dose_rp_col).value, 0.00252
        )

        # cleanup
        task.filename.delete()  # delete file so local testing doesn't get too messy!
        task.delete()  # not necessary, by hey, why not?

    def test_filters(self):
        """
        Tests that fluoro studies can be exported to XLSX  with single or multiple filters

        TODO: Add test study with no filter
        """
        filter_set = {"o": "-study_date"}
        pid = True
        name = False
        patient_id = True

        rfxlsx(filter_set, pid=pid, name=name, patid=patient_id, user=self.user)

        task = Exports.objects.all()[0]

        book = load_workbook(task.filename.path)
        philips_sheet = book["abdomen_2fps_25%"]
        siemens_sheet = book["fl_-_ang"]
        headers = siemens_sheet[1]

        filter_material_col = [
            i for i, x in enumerate(headers, start=1) if x.value == "Filter material"
        ][0]
        filter_thickness_col = [
            i
            for i, x in enumerate(headers, start=1)
            if x.value == "Mean filter thickness (mm)"
        ][0]
        time_col = [i for i, x in enumerate(headers, start=1) if x.value == "Time"][0]
        times = siemens_sheet[get_column_letter(time_col)]
        exp1_row = [
            i for i, x in enumerate(times, start=1) if x.value == "2016-05-12 10:11:54"
        ][0]

        self.assertEqual(
            philips_sheet.cell(row=2, column=filter_material_col).value, "Cu | Al"
        )
        self.assertEqual(
            philips_sheet.cell(row=2, column=filter_thickness_col).value,
            "0.1000 | 1.0000",
        )
        self.assertEqual(
            siemens_sheet.cell(row=exp1_row, column=filter_material_col).value, "Cu"
        )
        self.assertEqual(
            siemens_sheet.cell(row=exp1_row, column=filter_thickness_col).value,
            "0.6000",
        )

        # cleanup
        task.filename.delete()  # delete file so local testing doesn't get too messy!
        task.delete()  # not necessary, by hey, why not?

    def test_pulse_level_data(self):
        """Tests that RDSR with pulse level kVp, mA, pulse width data imports and exports with mean values"""
        filter_set = {"o": "-study_date"}
        pid = True
        name = False
        patient_id = True

        rfxlsx(filter_set, pid=pid, name=name, patid=patient_id, user=self.user)

        task = Exports.objects.all()[0]
        book = load_workbook(task.filename.path)

        eurocolumbus_sheet = book["vascular-knee-scopy-dose_level_"]
        eurocolumbus_headers = eurocolumbus_sheet[1]
        kvp_col = [
            i for i, x in enumerate(eurocolumbus_headers, start=1) if x.value == "kVp"
        ][0]
        ma_col = [
            i for i, x in enumerate(eurocolumbus_headers, start=1) if x.value == "mA"
        ][0]
        pulse_width_col = [
            i
            for i, x in enumerate(eurocolumbus_headers, start=1)
            if x.value == "Pulse width (ms)"
        ][0]
        exposure_time_col = [
            i for i, x in enumerate(eurocolumbus_headers, start=1) if x.value == "Time"
        ][0]
        target_row = 0
        for row_num in range(eurocolumbus_sheet.max_row):
            if (
                eurocolumbus_sheet.cell(row=row_num + 1, column=exposure_time_col).value
                == "2018-01-10 12:35:29"
            ):
                target_row = row_num + 1
                break

        self.assertAlmostEqual(
            eurocolumbus_sheet.cell(row=target_row, column=kvp_col).value,
            56.6666666666667,
        )
        self.assertAlmostEqual(
            eurocolumbus_sheet.cell(row=target_row, column=ma_col).value, 50.0
        )
        self.assertAlmostEqual(
            eurocolumbus_sheet.cell(row=target_row, column=pulse_width_col).value, 8.0
        )

        # cleanup
        task.filename.delete()  # delete file so local testing doesn't get too messy!
        task.delete()  # not necessary, by hey, why not?


class ExportRFArcadis(TransactionTestCase):
    def setUp(self):
        self.factory = RequestFactory()
        self.user = User.objects.create_user(
            username="jacob", email="jacob@…", password="top_secret"
        )
        eg = Group(name="exportgroup")
        eg.save()
        eg.user_set.add(self.user)
        eg.save()

        pid = PatientIDSettings.objects.create()
        pid.name_stored = True
        pid.name_hashed = False
        pid.id_stored = True
        pid.id_hashed = False
        pid.dob_stored = True
        pid.save()

        rf_siemens_arcadis = os.path.join("test_files", "RF-ESR-Siemens-Varic.dcm")
        root_tests = os.path.dirname(os.path.abspath(__file__))
        rdsr.rdsr(os.path.join(root_tests, rf_siemens_arcadis))

    def test_export_arcadis(self):
        filter_set = {"o": "-study_date"}
        pid = True
        name = False
        patient_id = True

        rfxlsx(filter_set, pid=pid, name=name, patid=patient_id, user=self.user)
